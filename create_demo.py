#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Tạo file Excel mẫu với công thức DocTien
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Tạo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Demo DocTien"
    
    # Tiêu đề
    ws['A1'] = 'SỐ TIỀN'
    ws['B1'] = 'BẰNG CHỮ (Sử dụng =DocTien(A2))'
    
    # Style cho header
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col in ['A1', 'B1']:
        ws[col].font = header_font
        ws[col].fill = header_fill
        ws[col].alignment = header_alignment
    
    # Dữ liệu mẫu
    data = [
        12345,
        1000000,
        2500.50,
        500000,
        1234567,
        0,
        999999999,
        15000,
        250000.75
    ]
    
    # Ghi dữ liệu vào cột A
    for idx, value in enumerate(data, start=2):
        cell = f'A{idx}'
        ws[cell] = value
        ws[cell].number_format = '#,##0.00'
        ws[cell].alignment = Alignment(horizontal="right")
    
    # Thêm công thức DocTien vào cột B
    for idx in range(2, 2 + len(data)):
        ws[f'B{idx}'] = f'=DocTien(A{idx})'
        ws[f'B{idx}'].alignment = Alignment(horizontal="left", vertical="center")
    
    # Điều chỉnh độ rộng cột
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 60
    ws.row_dimensions[1].height = 25
    
    # Border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws[f'A1:B{1 + len(data)}']:
        for cell in row:
            cell.border = thin_border
    
    # Thêm ghi chú
    ws['A13'] = 'LƯU Ý:'
    ws['A13'].font = Font(bold=True, size=11, color="DC2626")
    
    ws['A14'] = '1. Trước tiên, bạn cần cài đặt Add-in DocTien bằng cách chạy DocTien_Installer.exe'
    ws['A14'].font = Font(size=10)
    
    ws['A15'] = '2. Sau đó import file DocTien.bas vào Excel (xem file HUONG_DAN_SU_DUNG.txt)'
    ws['A15'].font = Font(size=10)
    
    ws['A16'] = '3. File này phải được lưu dạng .xlsm (Excel Macro-Enabled Workbook)'
    ws['A16'].font = Font(size=10)
    
    ws['A17'] = '4. Khi mở file, nhấn "Enable Content" để kích hoạt macro'
    ws['A17'].font = Font(size=10)
    
    # Lưu file vào thư mục hiện tại (linh hoạt cho mọi máy)
    output_file = os.path.join(os.path.dirname(__file__), 'Demo_DocTien.xlsx')
    wb.save(output_file)
    print(f"✅ Đã tạo file Demo_DocTien.xlsx thành công tại: {output_file}")
    
except ImportError:
    print("⚠️  Cần cài đặt openpyxl: pip install openpyxl")
    print("📝 Tạo file Excel đơn giản thay thế...")
    
    # Tạo CSV thay thế
    import csv
    
    output_file = os.path.join(os.path.dirname(__file__), 'Demo_DocTien.csv')
    with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['SỐ TIỀN', 'BẰNG CHỮ (Sử dụng =DocTien(A2))'])
        writer.writerow([12345, '=DocTien(A2)'])
        writer.writerow([1000000, '=DocTien(A3)'])
        writer.writerow([2500.50, '=DocTien(A4)'])
        writer.writerow([500000, '=DocTien(A5)'])
        writer.writerow([1234567, '=DocTien(A6)'])
        writer.writerow([0, '=DocTien(A7)'])
        writer.writerow([999999999, '=DocTien(A8)'])
    
    print(f"✅ Đã tạo file Demo_DocTien.csv thành công tại: {output_file}")
    print("💡 Mở file CSV bằng Excel, sau đó Save As → Excel Workbook (.xlsx)")
